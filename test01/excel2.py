'''
功能：用openpyxl读写excel
Created on 2019年1月21日
@author: Vostory
'''

# coding=utf-8
import openpyxl

filename = r'D:\\PracFolder\\Data\\111.xlsx'
wb = openpyxl.load_workbook(filename)  # 读文件
sheets = wb.sheetnames  # 获取读文件中所有的sheet，通过名字的方式
print('worksheets is %s' % sheets, type(sheets))

ws = wb[sheets[0]]  # 获取第一个sheet内容
# ws = workbook.sheet_by_name(u'Sheet1') # 通过工作表名称获取
# ws= workbook.sheets()[0] # 通过索引顺序获取
# ws= workbook.sheet_by_index(0) # 通过索引顺序获取
print(ws.title)  # 获得sheet名称

# 遍历sheet1中所有行row
num_rows = ws.max_row
print(num_rows)
for curr_row in range(num_rows):
    for cell1 in list(ws.rows)[curr_row]:
        print(cell1.value)
    print('row is %s' % curr_row)

# 遍历sheet1中所有列col
num_cols = ws.max_column
print(num_cols)
for curr_col in range(num_cols):
    for cell2 in list(ws.columns)[curr_col]:
        print(cell2.value)
    print('col is %s' % curr_col)

# 遍历sheet1中所有单元格cell
for rown in range(num_rows):  # 默认开始下标为0
    for coln in range(num_cols):  # 默认开始下标为0
        cell3 = ws.cell(rown + 1, coln + 1)  # 下标必须从1开始
        print(cell3)

print('数据读取结束！')

# 写入数据
outwb = openpyxl.Workbook()  # 打开一个将要写的excel文件
Sheet1 = outwb.create_sheet('Sheet1')  # 在将写的文件创建sheet,且命名为mySheet
print(outwb.sheetnames)  # 输出目前所有的工作表名称
ws = outwb.active  # 获取当前正在操作的表对象
ws.append(['电影名', '年份', '地区', '剧情类型', '导演', '主演', '评分', '评论人数', '简介'])
saveExcel = "D:\\PracFolder\\Data\\test.xlsx"
outwb.save(saveExcel)  # 一定要记得保存
print('数据写入结束！')

